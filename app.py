import io, re
from datetime import datetime
import streamlit as st
from openpyxl import load_workbook
from pdfminer.high_level import extract_text
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from pypdf import PdfReader, PdfWriter, Transformation
from pypdf._page import PageObject

# ---- Config ----
SIDE_MARGIN_MM = 2
TOP_MARGIN_MM = 4
STAMP_BOTTOM_MM = 12
INTER_GAP_MM = 1

BASE_CROP_L = 6
BASE_CROP_R = 6
BASE_CROP_T = 8
BASE_CROP_B = 8

LOW_TEXT_LINES = 4
SHORT_TEXT_CHARS = 80
EXTRA_CROP_LR = 14
EXTRA_CROP_T  = 18
EXTRA_CROP_B  = 28

def strip_diacritics(s):
    import unicodedata
    # ręczne poprawki dla ł/Ł
    s = s.replace("ł", "l").replace("Ł", "L")
    # standardowa normalizacja
    s_norm = unicodedata.normalize("NFD", s)
    s_clean = ''.join(c for c in s_norm if unicodedata.category(c) != "Mn")
    return s_clean


def read_excel_lookup(file_like):
    """
    Odczyt Excela:
      - wymagane kolumny: ZLECENIE, ilość palet, przewoźnik
      - opcjonalne kolumny: UWAGI, DOK
    Zwraca:
      lookup: numer -> (zlecenie, ilość palet, przewoźnik, uwagi, dok)
      all_nums: zbiór wszystkich numerów z kolumny ZLECENIE
    """
    wb = load_workbook(file_like, data_only=True); ws = wb.active
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        headers[str(v).strip().lower()] = col

    z_col   = headers.get("zlecenie")
    ilo_col = headers.get("ilośc palet") or headers.get("ilosc palet") or headers.get("ilość palet")
    pr_col  = headers.get("przewoźnik") or headers.get("przewoznik")
    uw_col  = headers.get("uwagi")   # opcjonalne
    dok_col = headers.get("dok")     # opcjonalne

    if not z_col or not ilo_col or not pr_col:
        raise ValueError("Excel musi mieć kolumny: ZLECENIE, ilość palet, przewoźnik (nagłówki w 1. wierszu).")

    lookup = {}
    all_nums = set()

    for row in range(2, ws.max_row + 1):
        z   = ws.cell(row=row, column=z_col).value
        il  = ws.cell(row=row, column=ilo_col).value
        pr  = ws.cell(row=row, column=pr_col).value
        uw  = ws.cell(row=row, column=uw_col).value if uw_col else None
        dok = ws.cell(row=row, column=dok_col).value if dok_col else None

        z   = "" if z   is None else str(z).strip()
        il  = "" if il  is None else str(il).strip()
        pr  = "" if pr  is None else str(pr).strip()
        uw  = "" if uw  is None else str(uw).strip()
        dok = "" if dok is None else str(dok).strip()

        parts = [p.strip() for p in re.split(r"[+;,/\s]+", z) if p.strip()]
        for p in parts:
            p2 = "".join(ch for ch in p if ch.isdigit())
            if p2.isdigit():
                all_nums.add(p2)
                lookup[p2] = (z, il, pr, uw, dok)

    return lookup, all_nums

NBSP = "\u00A0"; NNBSP = "\u202F"; THINSP = "\u2009"
def normalize_digits(s: str) -> str:
    import re
    return re.sub(r"[\s\-{}{}{}]".format(NBSP, NNBSP, THINSP), "", s)

def extract_candidates(text: str):
    import re
    normal = re.findall(r"\b\d{4,8}\b", text)
    fancy = re.findall(r"(?<!\d)(?:\d[\s\u00A0\u202F\u2009\-]?){4,9}(?!\d)", text)
    fancy = [normalize_digits(s) for s in fancy]
    so = [normalize_digits(m.group(1)) for m in re.finditer(r"Sales\s*[\r\n ]*Order[\s:]*([0-9\s\u00A0\u202F\u2009\-]{4,12})", text, flags=re.I)]
    cands = normal + fancy + so
    cands = [c for c in cands if c.isdigit() and 4 <= len(c) <= 8]
    out, seen = [], set()
    for c in cands:
        if c not in seen: out.append(c); seen.add(c)
    return out

def adaptive_crop_extra(text: str):
    lines = [ln for ln in (text or "").splitlines() if ln.strip()]
    sparse = (len(lines) <= LOW_TEXT_LINES) or (len((text or "")) < SHORT_TEXT_CHARS)
    if sparse: return (EXTRA_CROP_LR*mm, EXTRA_CROP_LR*mm, EXTRA_CROP_T*mm, EXTRA_CROP_B*mm)
    return (0,0,0,0)

def make_overlay(width, height, header, footer, uwagi="", dok="", font_size=12, margin_mm=8):
    """
    Nadruk w prawym dolnym rogu:
      - po PRAWEJ: nagłówek + stopka (ilość palet | przewoźnik)
      - po LEWEJ: DOK (nad) oraz UWAGI (pod)
    """
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(width, height))

    m = margin_mm * mm

    # --- prawa strona: nagłówek + stopka ---
    try:
        c.setFont("Helvetica-Bold", font_size)
    except Exception:
        c.setFont("Helvetica-Bold", font_size)

    # nagłówek
    c.drawRightString(width - m, m + font_size + 1, header)

    # stopka (ilość palet | przewoźnik)
    if footer:
        c.drawRightString(width - m, m, footer)

    # --- lewa strona: DOK nad UWAGAMI ---
    left_x = m
    y = m + font_size + 1

    # DOK (bez polskich znaków)
    if dok:
        try:
            c.setFont("Helvetica-Bold", font_size - 1)
        except Exception:
            c.setFont("Helvetica-Bold", font_size)
        dok_clean = strip_diacritics(dok)
        c.drawString(left_x, y, "DOK: {}".format(dok_clean))
        y -= (font_size + 2)

    # UWAGI (pogrubione, wielkie litery, bez polskich znaków)
    if uwagi:
        try:
            c.setFont("Helvetica-Bold", font_size - 1)
        except Exception:
            c.setFont("Helvetica-Bold", font_size)
        uw_clean = strip_diacritics(uwagi).upper()
        c.drawString(left_x, y, "UWAGI: {}".format(uw_clean))

    c.save()
    return buf.getvalue()


def make_summary_page(width, height, missing_from_pdf, missing_from_excel):
    """
    Raport końcowy:
    - pokazujemy tylko: ZLECENIA Z EXCELA NIEZNALEZIONE W PDF
    """
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(width, height))
    W, H = width, height
    try:
        c.setFont("Helvetica-Bold", 16)
    except Exception:
        c.setFont("Helvetica", 16)

    c.drawString(30, H-40, "RAPORT POROWNANIA DANYCH")
    y = H-80

    # sekcja: ZLECENIA Z EXCELA NIEZNALEZIONE W PDF
    c.setFont("Helvetica-Bold", 12)
    c.drawString(30, y, "ZLECENIA Z EXCELA NIEZNALEZIONE W PDF:")
    y -= 20

    c.setFont("Helvetica-Bold", 10)
    c.drawString(30, y, "ZLECENIE")
    y -= 12

    c.setLineWidth(0.5)
    c.line(30, y, W-30, y)
    y -= 10

    c.setFont("Helvetica", 10)
    if not missing_from_pdf:
        c.drawString(30, y, "(brak)")
        y -= 16
    else:
        for num in missing_from_pdf:
            c.drawString(30, y, str(num))
            y -= 14
            if y < 60:
                c.showPage()
                y = H-60
                c.setFont("Helvetica", 10)

    c.save()
    return buf.getvalue()

def annotate_pdf_web(pdf_bytes, xlsx_bytes, max_per_sheet):
    """
    Główna funkcja:
    - pomija pierwszą stronę PDF (zwykle parametry raportu),
    - dopasowuje strony do zleceń z Excela (oraz tworzy zlecenia "PDF-only"),
    - stempluje strony (ZLECENIE, ilość, przewoźnik, DOK, UWAGI),
    - na końcu dodaje stronę raportową z: ZLECENIA Z EXCELA NIEZNALEZIONE W PDF.
    """
    # Excel
    lookup, excel_numbers = read_excel_lookup(io.BytesIO(xlsx_bytes))

    # PDF wejściowy
    reader = PdfReader(io.BytesIO(pdf_bytes))

    groups: dict[str, list[int]] = {}
    page_meta: dict[int, tuple[str, str, str, str]] = {}
    page_text_cache: dict[int, str] = {}

    found_in_pdf: set[str] = set()
    pdf_candidates_all: set[str] = set()

    # --- przejście po stronach, pomijamy pierwszą (index 0) ---
    for i in range(1, len(reader.pages)):
        page_text = extract_text(io.BytesIO(pdf_bytes), page_numbers=[i]) or ""
        page_text_cache[i] = page_text

        cands = extract_candidates(page_text)

        # statystyka dopasowanych numerów
        for c in cands:
            if c in excel_numbers:
                found_in_pdf.add(c)
            else:
                pdf_candidates_all.add(c)

        picked_excel = next((n for n in cands if n in excel_numbers), None)
        picked_any = picked_excel or (cands[0] if cands else None)

        mapped = lookup.get(picked_excel) if picked_excel else None

        if mapped:
            # lookup: numer -> (zlecenie, ilość palet, przewoźnik, uwagi, dok)
            if len(mapped) == 5:
                z_full, il, pr, uw, dok = mapped
            else:
                z_full, il, pr = mapped
                uw = ""
                dok = ""

            key = z_full
            header = ("ZLECENIA (laczone): {}".format(strip_diacritics(z_full))
                      if "+" in z_full else "ZLECENIE: {}".format(strip_diacritics(z_full)))
            footer = "ilosc palet: {} | przewoznik: {}".format(strip_diacritics(il), strip_diacritics(pr))
        elif picked_any:
            key = picked_any
            header = "ZLECENIE: {}".format(picked_any)
            footer = "(brak danych w Excelu)"
            uw = ""
            dok = ""
        else:
            key = "_NO_ORDER_{}".format(i + 1)
            header = "(nie znaleziono numeru zlecenia na tej stronie)"
            footer = ""
            uw = ""
            dok = ""

        groups.setdefault(key, []).append(i)
        page_meta[i] = (header, footer, uw, dok)

    # --- sortowanie grup po numerach zleceń ---
    def key_sort(k: str):
        """
        Sortowanie grup:
        - NAJPIERW zlecenia nieznalezione w Excelu (PDF-only oraz _NO_ORDER_)
        - potem zlecenia powiązane z Excelem
        W każdej z tych dwóch grup sortujemy rosnąco po numerze zlecenia.
        """
        import re
        nums_str = re.findall(r"\d+", k)
        nums_int = [int(x) for x in nums_str] if nums_str else []
        # czy ten klucz ma numer występujący w Excelu?
        has_excel = any(n in excel_numbers for n in nums_str)
        group_flag = 1 if has_excel else 0  # 0 = PDF-only / NO_ORDER, 1 = z Excela
        primary = min(nums_int) if nums_int else 10**9
        return (group_flag, primary, k)

    ordered_keys = sorted(groups.keys(), key=key_sort)

    # --- parametry strony wynikowej ---
    W, H = A4
    margin_x = SIDE_MARGIN_MM * mm
    top_margin = TOP_MARGIN_MM * mm
    bot_stamp = STAMP_BOTTOM_MM * mm
    gap = INTER_GAP_MM * mm

    avail_w = W - 2 * margin_x
    avail_h = H - top_margin - bot_stamp

    base_crop_l = BASE_CROP_L * mm
    base_crop_r = BASE_CROP_R * mm
    base_crop_t = BASE_CROP_T * mm
    base_crop_b = BASE_CROP_B * mm

    writer = PdfWriter()
    writer.add_metadata({"/Producer": "Kersia PDF Stamper v1.6 (pypdf)"})

    # --- składanie stron na arkusze ---
    for gkey in ordered_keys:
        idxs = groups[gkey]

        for start in range(0, len(idxs), max_per_sheet):
            batch = idxs[start:start + max_per_sheet]

            items = []
            total_h = 0.0

            # najpierw policz skalowanie dla wszystkich stron w batchu
            for idx in batch:
                src = reader.pages[idx]
                sw = float(src.mediabox.width)
                sh = float(src.mediabox.height)

                ex_l, ex_r, ex_t, ex_b = adaptive_crop_extra(page_text_cache[idx])
                cl = base_crop_l + ex_l
                cr = base_crop_r + ex_r
                ct = base_crop_t + ex_t
                cb = base_crop_b + ex_b

                cw = max(10.0, sw - cl - cr)
                ch = max(10.0, sh - ct - cb)

                s = avail_w / cw
                dh = s * ch

                items.append((idx, cl, cr, ct, cb, s, dh))
                total_h += dh

            total_h += gap * max(0, len(batch) - 1)
            down = min(1.0, avail_h / total_h) if total_h > 0 else 1.0

            writer.add_blank_page(width=W, height=H)
            base_page = writer.pages[-1]

            y = H - top_margin

            for (idx, cl, cr, ct, cb, s, dh) in items:
                s *= down
                dh *= down

                x = margin_x - s * cl
                y2 = y - dh

                tmp = PageObject.create_blank_page(width=W, height=H)
                tmp.merge_page(reader.pages[idx])

                T = Transformation().translate(-cl, -cb).scale(s, s).translate(x, y2)
                tmp.add_transformation(T)
                base_page.merge_page(tmp)

                y = y2 - gap

            # nakładka z opisem (nagłówek, footer, DOK, UWAGI)
            ov = PdfReader(io.BytesIO(make_overlay(W, H, *page_meta[batch[0]])))
            base_page.merge_page(ov.pages[0])

    # --- raport: zlecenia z Excela nie znalezione w PDF ---
    excel_missing = sorted(list(excel_numbers - found_in_pdf), key=lambda x: int(x)) if excel_numbers else []

    if excel_missing:
        rep_pdf = PdfReader(io.BytesIO(make_summary_page(W, H, excel_missing, [])))
        for p in rep_pdf.pages:
            writer.add_page(p)

    # --- wynik ---
    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out.getvalue()



# ---- UI ----
st.set_page_config(page_title="Kersia PDF Stamper v1.6 (Raport)", page_icon="🧰", layout="centered")
st.title("Kersia — PDF Stamper (raport braków)")
excel_file = st.file_uploader("Plik Excel:", type=["xlsx", "xlsm", "xls"])
pdf_file = st.file_uploader("Plik PDF:", type=["pdf"])
max_per_sheet = st.slider("Maks. stron na kartkę", 1, 6, 3, 1)

if st.button("GENERUJ PDF", type="primary", disabled=not (excel_file and pdf_file)):
    try:
        data = annotate_pdf_web(pdf_file.read(), excel_file.read(), max_per_sheet)
        fname = "zlecenia_{}.pdf".format(datetime.now().strftime('%Y%m%d'))
        st.success("Gotowe! Pobierz poniżej.")
        st.download_button("Pobierz wynik", data=data, file_name=fname, mime="application/pdf")
    except Exception as e:
        st.error("Błąd: {}".format(repr(e)))
